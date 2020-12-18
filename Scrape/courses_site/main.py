import math
import asyncio
from datetime import datetime
import time

from bs4 import BeautifulSoup
from aiohttp import ClientSession

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class Scraper:
    def __init__(self):
        self.url = 'https://www.getwsodo.com/wp-admin/admin-ajax.php'
        self.post_per_request = 30
        self.total_pages = None

    async def get_pages(self):
        print("Looking how many pages I should parse...", end='')
        # Make a single request to get the total amount of pages
        content = await self.fetch_page(1)

        # I get the total amount of records, divide it per max items per page
        # and round up to get the total amount of pages.
        total_records = content['recordsTotal']
        total_pages = math.ceil(total_records / self.post_per_request)
        print(total_pages)
        return total_pages

    async def fetch_page(self, page_number):
        """ Asynchronously fetch pages """
        async with ClientSession() as session:
            async with session.post(self.url, data=self._generate_form(page_number)) as response:
                return await response.json()

    def _generate_form(self, page_number):
        """ Helper method that generates the form for the request with the needed page """
        return {
            'draw': page_number,
            'order[0][dir]': 'desc',
            'start': '0',
            'length': self.post_per_request,
            'table_id': 'ptp_9e3eebb6e40745fa_1',
            'action': 'ptp_load_posts'
        }

    @staticmethod
    def parse(pages_list):
        """ Parse the data from the list, getting title and date """
        print("Parsing pages...")
        data = []

        for page in pages_list:
            for element in page['data']:
                date = datetime.strptime(element['date'], "%B %d, %Y").date()
                title = BeautifulSoup(element['title'], 'html.parser').text
                data.append([date, title])

        return data

    @staticmethod
    def write_to_excel(courses):
        print("Writing to Excel file...")

        wb = Workbook()
        dst_filename = 'courses.xlsx'
        ws = wb.active
        ws.title = f"Courses from getWSOdownload"

        # Write headers as bold and centered
        cell_date = ws.cell(row=1, column=1, value="Date")
        cell_date.font = Font(name='Calibri', bold=True)
        cell_date.alignment = Alignment(horizontal='center')

        cell_title = ws.cell(row=1, column=2, value="Course Title")
        cell_title.font = Font(name='Calibri', bold=True)
        cell_title.alignment = Alignment(horizontal='center')

        # Write Date and Course Title
        for row, course in enumerate(courses, start=2):
            ws.cell(row=row, column=1).value = course[0]
            ws.cell(row=row, column=2).value = course[1]

        # Adjust column's widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60

        wb.save(dst_filename)


async def main():
    t0 = time.perf_counter()
    print("Starting scraping...")
    scraper = Scraper()
    total_pages = await scraper.get_pages()

    print("Requesting pages...")
    tasks = [asyncio.create_task(scraper.fetch_page(page)) for page in range(total_pages)]

    # Could use asyncio.gather() here and reduce to one line,
    # but this allows the program to give the user some feedback.
    all_pages = []
    for i, task in enumerate(tasks):
        print(f"\tRequesting page #{i + 1}")
        all_pages.append(await task)

    parsed = scraper.parse(all_pages)

    scraper.write_to_excel(parsed)
    print("\n\nScraping finished. Bye bye! :)\n")


if __name__ == '__main__':
    asyncio.run(main())
