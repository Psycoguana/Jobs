from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class PinnacleScraper:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        # Remove some annoying logs
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        # I use a downloaded chromedriver so the user doesn't need to download it.
        self.driver = webdriver.Chrome(executable_path='./chromedriver', options=chrome_options)

    def run(self):
        # We get the name of the game, and the link to the page where all it's bets reside.
        games = self.get_games_info(self.driver)
        matched_bets = self.get_matching_bets(games)
        self.driver.close()

        self.write_to_excel(matched_bets)

    def get_games_info(self, driver):
        games_url = 'https://www.pinnacle.com/en/football/nfl/matchups'
        print("Fetching website...")
        driver.get(games_url)

        # Wait until the page has fully loaded and the elements are visible.
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CLASS_NAME, "style_col__ncUvn")))
        print("Page loaded :)")

        # We get every available game.
        games = self.driver.find_elements_by_class_name('eventGameInfo')
        print(f"I found {len(games)} games to parse...\n")

        parsed_games = []
        for game in games:
            # Parse the name of the competing teams, also get the link to the bets page
            teams = game.find_elements_by_class_name('style_participantName__vRjBw')
            title = teams[0].text + " vs. " + teams[1].text
            parsed_games.append([title, game.get_attribute('href')])

        return parsed_games

    def get_matching_bets(self, matches):
        games = []

        for match_name, link in matches:
            print(f"Parsing {match_name}...")
            # This will return a list with every matching bet for a single game
            parsed_bets = self.find_receptions(link)
            games.append(Game(match_name, parsed_bets))

        return games

    def find_receptions(self, match_link):
        """ This method finds the bets that correspond to Receptions and Receiving Yds.
         Returns a list of the name, the over price and the under price. """

        bets = self.get_bets(match_link)

        parsed_bets = []
        for bet in bets:
            name = bet.find_element_by_class_name('style_titleText__jlbrV').text

            substrings = ["(Receptions)", "Receiving Yds"]
            if any(x in name for x in substrings):
                over, under = [x.text for x in bet.find_elements_by_class_name('label')]
                over = over.replace("ReceivingYards", "Receiving Yards").replace("PassReceptions", "Pass Receptions")
                under = under.replace("ReceivingYards", "Receiving Yards").replace("PassReceptions", "Pass Receptions")
                parsed_bets.append([name, over, under])

        return parsed_bets

    def get_bets(self, match_link):
        self.driver.get(match_link)

        print("\tWaiting for bets to load...")
        # Wait until the SEE ALL MARKETS button shows up, then click it.
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
            (By.XPATH, '//*[@id="root"]/div/div/div[2]/main/div/div[2]/div[2]/button'))).click()

        bets_class = 'style_primary__awRGO'
        # Wait until all bets are loaded.
        WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.CLASS_NAME, bets_class))).click()
        print("\tBets loaded\n")

        bets = self.driver.find_elements_by_class_name(bets_class)
        return bets

    def write_to_excel(self, games):
        print("Writing to Excel file...")

        wb = Workbook()
        dst_filename = 'pinnacle_nfl.xlsx'

        for game in games:
            # Some games won't have the bets we need, I check this so we don't create an empty sheet.
            if game.bets:
                # Titles longer than 31 characters are not recommended
                current_ws = wb.create_sheet(game.game_name[:31])
                self.format_excel_header(current_ws)

                for row, bet in enumerate(game.bets, start=2):
                    current_ws.cell(row=row, column=1).value = bet[0]
                    current_ws.cell(row=row, column=2).value = bet[1]
                    current_ws.cell(row=row, column=3).value = bet[2]

        # OpenPyxl creates a default sheet which we don't really use, so I just remove it once the data is written.
        wb.remove(wb['Sheet'])
        wb.save(dst_filename)

    @staticmethod
    def format_excel_header(ws):
        # Write headers
        ws.cell(row=1, column=1, value="Player")
        ws.cell(row=1, column=2, value="Over")
        ws.cell(row=1, column=3, value="Under")

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40

        # Format as bold and centered.
        for i in range(1, 4):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(name='Calibri', bold=True)
            cell.alignment = Alignment(horizontal='center')


class Game:
    """ This class purely holds data. Would have used a dataclass, but that's only available from python 3.9 """

    def __init__(self, game_name, bets):
        self.game_name = game_name
        self.bets = bets


if __name__ == '__main__':
    print("Starting scraper...")
    scraper = PinnacleScraper()
    scraper.run()
    print("\n\nThe script has finished. You can close this window now :)")
