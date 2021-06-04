import datetime
import json
import logging
import math
import re
from typing import List

import requests
import telegram
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from telegram import Update
from telegram.ext import Updater, CommandHandler, CallbackContext, MessageHandler, Filters

from Database import Database
from House import House


class Bot:
    def __init__(self):
        logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S',
                            level=logging.INFO)
        logging.info("Starting bot...")
        self.token = ''
        self.database = Database()

    def main(self):
        updater = Updater(self.token)
        dispatcher = updater.dispatcher

        # Add command handlers.
        dispatcher.add_handler(CommandHandler("start", self.start))
        dispatcher.add_handler(CommandHandler("del", self.remove_link))
        dispatcher.add_handler(MessageHandler(Filters.text & ~Filters.command, self.handle_new_links))

        # Start the Bot
        updater.start_polling()
        updater.idle()

    def start(self, update: Update, _: CallbackContext):
        message = "Hey there!\n" \
                  "To add a house just send a bunch of link.\n" \
                  "I recommend you do it at once or else you'll get a bunch of excel files.\n\n" \
                  "To remove them type /del house_link"

        update.message.reply_text(message)

    def remove_link(self, update: Update, _: CallbackContext):
        # Remove command from message.
        link = update.message.text.replace('/del ', '')

        if link.startswith('https://'):
            self.database.remove(link)
            logging.info(f"Removing House: {link}")
            update.message.reply_text(f'House Removed:\n{link}', disable_web_page_preview=True)

    def handle_new_links(self, update: Update, context: CallbackContext):
        # Send typing action.
        context.bot.send_chat_action(chat_id=update.message.chat_id, action=telegram.ChatAction.TYPING)
        messages = self.parse_message(update.message.text)

        houses = []
        for message in messages:
            if message.startswith('https://'):
                try:
                    logging.debug(f"Scraping: {message}")
                    new_house = self.get_house_details(url=message)
                    houses.append(new_house)
                except Exception as error:
                    error_message = f"Error while trying to scrape.\n{error}"
                    logging.error(error_message)
                    update.message.reply_text(error_message)
            else:
                error_message = f"Link not recognized -> {message}"
                logging.info(error_message)
                update.message.reply_text(error_message)

        # If no house was correctly scraped, return so we don't try adding anything to the database,
        # and neither send a new excel file.
        if not houses:
            return

        # Add new data
        self.database.insert_new_values(houses)

        db_houses = self.database.get_values()

        # If the database is not empty, unpack it's values, generate an excel file and send it
        if db_houses:
            # Unpack all houses into House objects and store them in a list.
            all_houses = [House(*x) for x in db_houses]

            excel_file = self.generate_excel(all_houses)
            self.send_excel(context=context, chat_id=update.message.chat_id, excel_filename=excel_file)

    def parse_message(self, message: str):
        parsed_messages = []

        messages = message.split('\n')
        parsed_messages.extend(messages)

        # Don't return None values.
        return [message.lower() for message in parsed_messages if message]

    def fetch_house_data(self, url):
        headers = {
            'authority': 'REDACTED',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
        }

        api_link = 'REDACTED'
        property_id = self.get_property_id(url)
        params = {
            "propertyId": property_id,
            'accessLevel': '1',
        }

        # First chars are "{}&&", which we need to remove to get the actual JSON we want.
        response = requests.get(api_link, headers=headers, params=params).content[4:]

        return json.loads(response)['payload']

    def get_house_details(self, url):
        house_data = self.fetch_house_data(url)

        public_records = house_data['publicRecordsInfo']
        history_info = house_data['propertyHistoryInfo']

        address_info = public_records['addressInfo']
        street = address_info['street']
        city = address_info['city']
        state = address_info['state']
        zip = address_info['zip']
        address = ', '.join([street, city, state, zip])

        year_built = public_records['basicInfo']['yearBuilt']
        try:
            taxableLandValue, taxableImprovementValue, rollYear, taxesDue = [x for x in
                                                                             public_records['taxInfo'].values()]
        except Exception:
            taxableLandValue, taxableImprovementValue, rollYear, taxesDue = None, None, None, None

        price = None
        time_on_market = None

        events = history_info['events']
        for event in events:

            if not time_on_market:
                if 'Listed' in event['eventDescription']:
                    # Truncate last 3 digits because they're milliseconds
                    listed_on_epoch = str(event['eventDate'])[0: 10]
                    listed_on_epoch = int(listed_on_epoch)
                    time_on_market = datetime.datetime.fromtimestamp(listed_on_epoch).strftime("%Y-%m-%d")

            if not price:
                try:
                    price = float(event["price"])
                except KeyError:
                    # This event doesn't have a price, ignore it.
                    continue

        square_feet = public_records['basicInfo']['totalSqFt']
        price_per_square_feet = math.floor(price / square_feet)

        status = ''
        event_description = events[0]['eventDescription']
        if event_description == 'Contingent' or 'sold' in event_description.lower():
            status = "SOLD"

        highschools_list = house_data['schoolsAndDistrictsInfo']['highSchools']
        highschools_list = [x['name'] for x in highschools_list]

        highschools = ', '.join(highschools_list)

        return House(url,address, city, status, year_built, time_on_market, f'${price}', square_feet,
                     f'${price_per_square_feet}',
                     highschools, taxableLandValue, taxableImprovementValue, rollYear, taxesDue)

    def get_property_id(self, link):
        # Extract the property id from the link, in order to use it later to build the api request.
        pattern = '\/home\/([0-9]+)'

        return re.search(pattern, link).group(1)

    def generate_excel(self, houses: List[House]):
        logging.info("Generating Excel file...")

        wb = Workbook()
        dst_filename = 'Houses.xlsx'
        ws = wb.active
        ws.title = f"Houses"

        # Write headers as bold and centered
        headers = ["Address", "City", "Price", "Status", "Year Built", "Listed on Market", "Square Footage",
                   "Price Per Square Feet",
                   "Schools", "Taxable Land Value", "Taxable Improvement Value", "Roll Year", "Taxed Due", "Link"]
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=column)
            cell.value = header
            cell.font = Font(name='Calibri', bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write Date and Course Title
        for row, house in enumerate(houses, start=2):
            ws.cell(row=row, column=1).value = house.address
            ws.cell(row=row, column=2).value = house.city
            ws.cell(row=row, column=3).value = house.price
            ws.cell(row=row, column=4).value = house.status
            ws.cell(row=row, column=5).value = house.property_year
            ws.cell(row=row, column=6).value = house.date_on_market
            ws.cell(row=row, column=7).value = house.square_feet
            ws.cell(row=row, column=8).value = house.price_per_square_feet
            ws.cell(row=row, column=9).value = house.school_names
            ws.cell(row=row, column=10).value = house.taxableLandValue
            ws.cell(row=row, column=11).value = house.taxableImprovementValue
            ws.cell(row=row, column=12).value = house.rollYear
            ws.cell(row=row, column=13).value = house.taxesDue
            ws.cell(row=row, column=14, value='LINK').hyperlink = house.link

        # Adjust column's widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 19
        ws.column_dimensions['I'].width = 20

        wb.save(dst_filename)
        return dst_filename

    def send_excel(self, context, chat_id, excel_filename):
        logging.info("Sending Excel file...")
        with open(f'./{excel_filename}', 'rb') as excel_file:
            context.bot.send_document(chat_id=chat_id, caption="Here's your file!", document=excel_file)


if __name__ == '__main__':
    try:
        bot = Bot()
        bot.main()
    except Exception as error:
        logging.error(f'Error while running the bot:\n{error}')
