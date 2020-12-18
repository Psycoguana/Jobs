import re
import csv
from typing import List
from datetime import datetime
from configparser import ConfigParser

import praw
from praw.models import ListingGenerator, Comment
from praw.models.comment_forest import CommentForest

from openpyxl import Workbook


class CommentSearcher:
    def __init__(self):
        user_agent = 'CommentSearcher'
        config = ConfigParser()
        config.read('config.ini')
        self.subreddit = config['DEFAULT']['SUBREDDIT']
        self.querys = config['DEFAULT']['QUERYS'].replace(' ', '').split(',')

        self.reddit = praw.Reddit(
            client_id=config['DEFAULT']['CLIENT_ID'],
            client_secret=config['DEFAULT']['CLIENT_SECRET'],
            user_agent=user_agent
        )

    def get_comment_forests(self) -> List[CommentForest]:
        """
        Reddit allows praw to get comments matching a certain keyword, but returns a CommentForest
        (parent comments). This method returns a list of CommentForest objects for later parsing.
        :return: List of CommentForest objects
        """
        comment_forests_list: List[CommentForest] = []
        subreddit = self.reddit.subreddit(self.subreddit)

        for query in self.querys:
            print(f"Searching {query} in r/{self.subreddit} (this may take awhile)")
            matching_submissions: ListingGenerator = subreddit.search(query=query, sort='comments')
            for submission in matching_submissions:
                print(f"Resolving submission {submission}", end='\r')
                comment_forests_list.append(submission.comments)

        return comment_forests_list

    def get_comment_objects(self) -> List[Comment]:
        comments_forests: List[CommentForest] = self.get_comment_forests()
        comment_objects_list = []

        for comment_forest in comments_forests:
            # Resolve child comments so they can be parsed. If this is not done, the CommentForest object might contain
            # MoreComments objects instead of just Comment objects.
            comment_forest.replace_more(limit=None)

            for comment_object in comment_forest.list():
                print(f"Resolving comment {comment_object}", end='\r')
                comment_objects_list.append(comment_object)

        return comment_objects_list

    def search_in_comments(self, comments: List[Comment]):
        matched_comments: [Comment] = []

        for comment in comments:
            for query in self.querys:
                # This regex should match the query with the comment's body
                # as long as the latter has a space or nothing behind and/or ahead.
                if re.search(r"((?:(?<= )|^){}(?= |$))".format(re.escape(query)), comment.body, re.IGNORECASE):
                    matched_comments.append(comment)

        return matched_comments

    @staticmethod
    def write_to_csv(comments):
        print("Writing to csv file...")

        with open('comments.csv', 'w', newline='') as csv_file:
            fieldnames = ['timestamp', 'score', 'text']
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()

            for comment in comments:
                writer.writerow(
                    {'timestamp': datetime.fromtimestamp(comment.created_utc).strftime('%Y-%m-%d %H:%M:%S'),
                     'score': comment.score,
                     'text': comment.body}
                )

    def write_to_excel(self, comments):
        print("Writing to Excel file...")

        wb = Workbook()
        dst_filename = 'comments.xlsx'
        ws = wb.active
        ws.title = f"Comments from {self.subreddit}"

        # Write headers
        ws.cell(row=1, column=1, value="Timestamp")
        ws.cell(row=1, column=2, value="Score")
        ws.cell(row=1, column=3, value="Text")

        # Write timestamp, score and text
        max_comment_len = 1
        for row, comment in enumerate(comments, start=2):
            # Save the maximum possible length for a comment so the column can be adjusted accordingly.
            max_comment_len = max_comment_len if (max_comment_len > len(comment.body)) else len(comment.body)

            formatted_date = datetime.fromtimestamp(comment.created_utc).strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row, column=1).value = formatted_date
            ws.cell(row=row, column=2).value = comment.score
            ws.cell(row=row, column=3).value = comment.body

        # Adjust column's widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        # Limit the width to 60 so it's each row is not an amazingly long line of text
        ws.column_dimensions['C'].width = max_comment_len if max_comment_len < 60 else 60

        wb.save(dst_filename)


if __name__ == '__main__':
    app = CommentSearcher()
    comment_objects = app.get_comment_objects()
    matching_comments = app.search_in_comments(comments=comment_objects)

    app.write_to_csv(comments=matching_comments)
    app.write_to_excel(comments=matching_comments)
